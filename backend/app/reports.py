"""Rapports financiers (PDF) et export des transactions (CSV) pour les financiers de la LECIM."""

import csv
import datetime
import io
from pathlib import Path

from fpdf import FPDF
from fpdf.fonts import FontFace
from sqlalchemy.orm import Session, joinedload

from . import models
from .finances_constants import RECETTE_CATEGORIES

LOGO_PATH = Path(__file__).resolve().parent / "static" / "img" / "logo.jpg"


def _add_logo(pdf: FPDF) -> None:
    if LOGO_PATH.exists():
        pdf.image(str(LOGO_PATH), x=pdf.w - pdf.r_margin - 22, y=8, w=22, h=22)


def current_annee_scolaire(today: datetime.date | None = None) -> str:
    today = today or datetime.date.today()
    if today.month >= 9:
        return f"{today.year}-{today.year + 1}"
    return f"{today.year - 1}-{today.year}"


def money(value: int) -> str:
    return f"{value:,}".replace(",", " ") + " FCFA"


_PDF_CHAR_REPLACEMENTS = {
    "—": "-",  # em dash
    "–": "-",  # en dash
    "‘": "'",
    "’": "'",
    "“": '"',
    "”": '"',
    "…": "...",
    "•": "-",
}


def pdf_safe(text: str | None) -> str:
    """Remplace les caractères typographiques hors Latin-1 (tirets cadratins,
    guillemets courbes, etc.) que la police Helvetica de base ne sait pas rendre."""
    if not text:
        return ""
    for char, replacement in _PDF_CHAR_REPLACEMENTS.items():
        text = text.replace(char, replacement)
    return text


def etablissements_en_retard(db: Session, annee_scolaire: str | None = None) -> list[dict]:
    """Établissements dont la cotisation de l'année scolaire n'est pas soldée."""
    annee_scolaire = annee_scolaire or current_annee_scolaire()
    etablissements = db.query(models.Etablissement).order_by(models.Etablissement.nom).all()
    cotisations = {
        c.etablissement_id: c
        for c in db.query(models.Cotisation)
        .filter(models.Cotisation.annee_scolaire == annee_scolaire)
        .all()
    }
    en_retard = []
    for e in etablissements:
        cotisation = cotisations.get(e.id)
        if cotisation is None:
            from .finances_constants import cotisation_rule

            rule = cotisation_rule(e.statut)
            en_retard.append(
                {
                    "etablissement": e,
                    "montant_du": rule["montant_du"],
                    "montant_paye": 0,
                    "reste": rule["montant_du"],
                }
            )
        elif cotisation.statut_paiement != "paye":
            en_retard.append(
                {
                    "etablissement": e,
                    "montant_du": cotisation.montant_du,
                    "montant_paye": cotisation.montant_paye,
                    "reste": cotisation.montant_du - cotisation.montant_paye,
                }
            )
    return en_retard


def multi_year_financial_summary(db: Session) -> list[dict]:
    """Totaux financiers regroupés par année scolaire, toutes années confondues,
    pour visualiser l'évolution pluriannuelle des finances de la LECIM."""
    buckets: dict[str, dict] = {}

    def bucket(annee: str) -> dict:
        return buckets.setdefault(
            annee,
            {"adhesions": 0, "cotisations": 0, "recettes": 0, "ventes_livres": 0, "droits_examens": 0, "depenses": 0},
        )

    for a in db.query(models.Adhesion).all():
        bucket(current_annee_scolaire(a.date_paiement))["adhesions"] += a.montant
    for c in db.query(models.Cotisation).filter(models.Cotisation.date_paiement.isnot(None)).all():
        bucket(c.annee_scolaire)["cotisations"] += c.montant_paye
    for r in db.query(models.Recette).all():
        bucket(current_annee_scolaire(r.date))["recettes"] += r.montant
    for v in db.query(models.VenteLivre).all():
        bucket(current_annee_scolaire(v.date))["ventes_livres"] += v.montant
    for de in db.query(models.DroitExamen).all():
        bucket(current_annee_scolaire(de.date))["droits_examens"] += de.montant
    for d in db.query(models.Depense).all():
        bucket(current_annee_scolaire(d.date))["depenses"] += d.montant

    result = []
    for annee in sorted(buckets.keys()):
        b = buckets[annee]
        total_entrees = b["adhesions"] + b["cotisations"] + b["recettes"] + b["ventes_livres"] + b["droits_examens"]
        result.append(
            {
                "annee": annee,
                "adhesions": b["adhesions"],
                "cotisations": b["cotisations"],
                "recettes": b["recettes"],
                "ventes_livres": b["ventes_livres"],
                "droits_examens": b["droits_examens"],
                "depenses": b["depenses"],
                "total_entrees": total_entrees,
                "solde": total_entrees - b["depenses"],
            }
        )
    return result


def _period_data(db: Session, date_debut: datetime.date, date_fin: datetime.date) -> dict:
    adhesions = (
        db.query(models.Adhesion)
        .options(joinedload(models.Adhesion.etablissement))
        .filter(models.Adhesion.date_paiement >= date_debut, models.Adhesion.date_paiement <= date_fin)
        .all()
    )
    cotisations = (
        db.query(models.Cotisation)
        .options(joinedload(models.Cotisation.etablissement))
        .filter(
            models.Cotisation.date_paiement.isnot(None),
            models.Cotisation.date_paiement >= date_debut,
            models.Cotisation.date_paiement <= date_fin,
        )
        .all()
    )
    recettes = (
        db.query(models.Recette)
        .filter(models.Recette.date >= date_debut, models.Recette.date <= date_fin)
        .all()
    )
    ventes_livres = (
        db.query(models.VenteLivre)
        .filter(models.VenteLivre.date >= date_debut, models.VenteLivre.date <= date_fin)
        .all()
    )
    droits_examens = (
        db.query(models.DroitExamen)
        .options(joinedload(models.DroitExamen.etablissement))
        .filter(models.DroitExamen.date >= date_debut, models.DroitExamen.date <= date_fin)
        .all()
    )
    depenses = (
        db.query(models.Depense)
        .filter(models.Depense.date >= date_debut, models.Depense.date <= date_fin)
        .all()
    )

    total_adhesions = sum(a.montant for a in adhesions)
    total_cotisations = sum(c.montant_paye for c in cotisations)
    total_recettes = sum(r.montant for r in recettes)
    total_ventes_livres = sum(v.montant for v in ventes_livres)
    total_droits_examens = sum(d.montant for d in droits_examens)
    total_depenses = sum(d.montant for d in depenses)

    recettes_par_categorie: dict[str, int] = {}
    for r in recettes:
        recettes_par_categorie[r.categorie] = recettes_par_categorie.get(r.categorie, 0) + r.montant

    total_entrees = total_adhesions + total_cotisations + total_recettes + total_ventes_livres + total_droits_examens

    return {
        "adhesions": adhesions,
        "cotisations": cotisations,
        "recettes": recettes,
        "ventes_livres": ventes_livres,
        "droits_examens": droits_examens,
        "depenses": depenses,
        "total_adhesions": total_adhesions,
        "total_cotisations": total_cotisations,
        "total_recettes": total_recettes,
        "total_ventes_livres": total_ventes_livres,
        "total_droits_examens": total_droits_examens,
        "total_depenses": total_depenses,
        "recettes_par_categorie": recettes_par_categorie,
        "total_entrees": total_entrees,
        "solde_periode": total_entrees - total_depenses,
    }


def _solde_cumule_au(db: Session, date_fin: datetime.date) -> int:
    total_adhesions = sum(
        a.montant for a in db.query(models.Adhesion).filter(models.Adhesion.date_paiement <= date_fin)
    )
    total_cotisations = sum(
        c.montant_paye
        for c in db.query(models.Cotisation).filter(
            models.Cotisation.date_paiement.isnot(None), models.Cotisation.date_paiement <= date_fin
        )
    )
    total_recettes = sum(r.montant for r in db.query(models.Recette).filter(models.Recette.date <= date_fin))
    total_ventes_livres = sum(
        v.montant for v in db.query(models.VenteLivre).filter(models.VenteLivre.date <= date_fin)
    )
    total_droits_examens = sum(
        d.montant for d in db.query(models.DroitExamen).filter(models.DroitExamen.date <= date_fin)
    )
    total_depenses = sum(d.montant for d in db.query(models.Depense).filter(models.Depense.date <= date_fin))
    return (
        total_adhesions + total_cotisations + total_recettes + total_ventes_livres + total_droits_examens
        - total_depenses
    )


GREEN = (15, 122, 76)
GOLD = (224, 165, 44)
GRAY = (170, 170, 170)


def _short_money(value: int) -> str:
    if abs(value) >= 1_000_000:
        return f"{value / 1_000_000:.1f}M"
    if abs(value) >= 1_000:
        return f"{value / 1_000:.0f}k"
    return str(value)


def _draw_bar_chart(pdf: FPDF, x: float, y: float, width: float, height: float, bars: list[tuple]) -> None:
    """Dessine un histogramme simple. bars: liste de (label, valeur, couleur_rgb)."""
    if not bars:
        return
    max_val = max((abs(b[1]) for b in bars), default=0) or 1
    n = len(bars)
    gap = 6
    bar_width = (width - gap * (n - 1)) / n

    pdf.set_draw_color(*GRAY)
    pdf.line(x, y + height, x + width, y + height)

    for i, (label, value, color) in enumerate(bars):
        bar_h = (abs(value) / max_val) * (height - 8)
        bx = x + i * (bar_width + gap)
        by = y + height - bar_h
        pdf.set_fill_color(*color)
        pdf.rect(bx, by, bar_width, bar_h, style="F")

        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(30, 30, 30)
        pdf.set_xy(bx - 5, by - 5)
        pdf.cell(bar_width + 10, 5, _short_money(value), align="C")

        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(80, 80, 80)
        pdf.set_xy(bx - 5, y + height + 1.5)
        pdf.cell(bar_width + 10, 5, label, align="C")

    pdf.set_text_color(30, 30, 30)
    pdf.set_fill_color(255, 255, 255)
    pdf.set_draw_color(0, 0, 0)
    pdf.set_xy(x, y + height + 8)


def _draw_grouped_bar_chart(
    pdf: FPDF,
    x: float,
    y: float,
    width: float,
    height: float,
    groups: list[str],
    serie_a: list[int],
    serie_b: list[int],
    label_a: str,
    label_b: str,
) -> None:
    """Histogramme groupé à deux séries (comparaison de deux périodes)."""
    n = len(groups)
    max_val = max([abs(v) for v in serie_a + serie_b] or [0]) or 1
    group_gap = 10
    group_width = (width - group_gap * (n - 1)) / n
    bar_gap = 2
    bar_width = (group_width - bar_gap) / 2

    pdf.set_draw_color(*GRAY)
    pdf.line(x, y + height, x + width, y + height)

    for i, label in enumerate(groups):
        gx = x + i * (group_width + group_gap)
        for j, (value, color) in enumerate(((serie_a[i], GREEN), (serie_b[i], GOLD))):
            bar_h = (abs(value) / max_val) * (height - 8)
            bx = gx + j * (bar_width + bar_gap)
            by = y + height - bar_h
            pdf.set_fill_color(*color)
            pdf.rect(bx, by, bar_width, bar_h, style="F")
            pdf.set_font("Helvetica", "B", 7)
            pdf.set_text_color(30, 30, 30)
            pdf.set_xy(bx - 6, by - 5)
            pdf.cell(bar_width + 12, 5, _short_money(value), align="C")

        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(80, 80, 80)
        pdf.set_xy(gx, y + height + 1.5)
        pdf.cell(group_width, 5, label, align="C")

    legend_y = y + height + 9
    pdf.set_fill_color(*GREEN)
    pdf.rect(x, legend_y, 4, 4, style="F")
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(30, 30, 30)
    pdf.set_xy(x + 6, legend_y - 1.5)
    pdf.cell(60, 6, label_a)

    pdf.set_fill_color(*GOLD)
    pdf.rect(x + 70, legend_y, 4, 4, style="F")
    pdf.set_xy(x + 76, legend_y - 1.5)
    pdf.cell(60, 6, label_b)

    pdf.set_fill_color(255, 255, 255)
    pdf.set_draw_color(0, 0, 0)
    pdf.set_text_color(30, 30, 30)
    pdf.set_xy(x, legend_y + 8)


def generate_financial_report_pdf(
    db: Session, date_debut: datetime.date, date_fin: datetime.date, generated_by: str
) -> bytes:
    data = _period_data(db, date_debut, date_fin)
    solde_cumule = _solde_cumule_au(db, date_fin)
    retards = etablissements_en_retard(db)

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()
    _add_logo(pdf)

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(11, 61, 46)
    pdf.cell(0, 10, "LECIM - Rapport financier", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(80, 80, 80)
    periode = f"Periode du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
    pdf.cell(0, 7, periode, new_x="LMARGIN", new_y="NEXT")
    pdf.cell(
        0, 7, f"Genere le {datetime.date.today().strftime('%d/%m/%Y')} par {generated_by}",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.ln(4)

    def section_title(text: str) -> None:
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(11, 61, 46)
        pdf.cell(0, 9, text, new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(30, 30, 30)

    def kv_row(label: str, value: str, bold: bool = False) -> None:
        pdf.set_font("Helvetica", "B" if bold else "", 11)
        pdf.cell(100, 7, label)
        pdf.cell(0, 7, value, new_x="LMARGIN", new_y="NEXT")

    section_title("Resume de la periode")
    kv_row("Droits d'adhesion encaisses", money(data["total_adhesions"]))
    kv_row("Cotisations encaissees", money(data["total_cotisations"]))
    kv_row("Autres recettes", money(data["total_recettes"]))
    kv_row("Ventes de livres", money(data["total_ventes_livres"]))
    kv_row("Droits d'examens", money(data["total_droits_examens"]))
    kv_row("Total des entrees", money(data["total_entrees"]), bold=True)
    kv_row("Depenses", money(data["total_depenses"]))
    kv_row("Solde de la periode", money(data["solde_periode"]), bold=True)
    pdf.ln(2)
    kv_row("Solde cumule au " + date_fin.strftime("%d/%m/%Y"), money(solde_cumule), bold=True)
    pdf.ln(10)

    _draw_bar_chart(
        pdf,
        pdf.l_margin,
        pdf.get_y(),
        pdf.w - pdf.l_margin - pdf.r_margin,
        40,
        [
            ("Entrees", data["total_entrees"], GREEN),
            ("Depenses", data["total_depenses"], GOLD),
            ("Solde periode", data["solde_periode"], GREEN if data["solde_periode"] >= 0 else (192, 57, 43)),
        ],
    )
    pdf.ln(4)

    if data["recettes_par_categorie"]:
        section_title("Autres recettes par categorie")
        chart_bars = [
            (RECETTE_CATEGORIES.get(cat, cat)[:14], montant, GOLD)
            for cat, montant in data["recettes_par_categorie"].items()
        ]
        _draw_bar_chart(pdf, pdf.l_margin, pdf.get_y(), pdf.w - pdf.l_margin - pdf.r_margin, 36, chart_bars)
        pdf.ln(4)
        pdf.set_font("Helvetica", "", 10)
        with pdf.table(col_widths=(110, 60), text_align=("LEFT", "RIGHT")) as table:
            row = table.row()
            row.cell("Categorie", style=FontFace(emphasis="BOLD"))
            row.cell("Montant", style=FontFace(emphasis="BOLD"))
            for categorie, montant in data["recettes_par_categorie"].items():
                row = table.row()
                row.cell(RECETTE_CATEGORIES.get(categorie, categorie))
                row.cell(money(montant))
        pdf.ln(4)

    if data["ventes_livres"]:
        section_title("Ventes de livres sur la periode")
        pdf.set_font("Helvetica", "", 10)
        with pdf.table(col_widths=(40, 90, 30, 40), text_align=("LEFT", "LEFT", "RIGHT", "RIGHT")) as table:
            row = table.row()
            for h in ("Date", "Titre", "Quantite", "Montant"):
                row.cell(h, style=FontFace(emphasis="BOLD"))
            for v in data["ventes_livres"]:
                row = table.row()
                row.cell(v.date.strftime("%d/%m/%Y"))
                row.cell(pdf_safe(v.titre))
                row.cell(str(v.quantite))
                row.cell(money(v.montant))
        pdf.ln(4)

    if data["droits_examens"]:
        section_title("Droits d'examens sur la periode")
        pdf.set_font("Helvetica", "", 10)
        with pdf.table(col_widths=(35, 60, 45, 40), text_align=("LEFT", "LEFT", "LEFT", "RIGHT")) as table:
            row = table.row()
            for h in ("Date", "Libelle", "Etablissement", "Montant"):
                row.cell(h, style=FontFace(emphasis="BOLD"))
            for de in data["droits_examens"]:
                row = table.row()
                row.cell(de.date.strftime("%d/%m/%Y"))
                row.cell(pdf_safe(de.libelle))
                row.cell(pdf_safe(de.etablissement.nom) if de.etablissement else "-")
                row.cell(money(de.montant))
        pdf.ln(4)

    if data["cotisations"]:
        section_title("Cotisations encaissees sur la periode")
        pdf.set_font("Helvetica", "", 10)
        with pdf.table(col_widths=(70, 40, 40, 40), text_align=("LEFT", "LEFT", "RIGHT", "RIGHT")) as table:
            row = table.row()
            for h in ("Etablissement", "Annee scolaire", "Verse", "Part bureau local"):
                row.cell(h, style=FontFace(emphasis="BOLD"))
            for c in data["cotisations"]:
                row = table.row()
                row.cell(pdf_safe(c.etablissement.nom))
                row.cell(c.annee_scolaire)
                row.cell(money(c.montant_paye))
                row.cell(money(c.part_bureau_local))
        pdf.ln(4)

    if data["depenses"]:
        section_title("Depenses de la periode")
        pdf.set_font("Helvetica", "", 10)
        with pdf.table(col_widths=(40, 100, 50), text_align=("LEFT", "LEFT", "RIGHT")) as table:
            row = table.row()
            for h in ("Date", "Libelle", "Montant"):
                row.cell(h, style=FontFace(emphasis="BOLD"))
            for d in data["depenses"]:
                row = table.row()
                row.cell(d.date.strftime("%d/%m/%Y"))
                row.cell(pdf_safe(d.libelle))
                row.cell(money(d.montant))
        pdf.ln(4)

    if retards:
        section_title(f"Etablissements en retard de cotisation ({current_annee_scolaire()})")
        pdf.set_font("Helvetica", "", 10)
        with pdf.table(col_widths=(70, 40, 40, 40), text_align=("LEFT", "RIGHT", "RIGHT", "RIGHT")) as table:
            row = table.row()
            for h in ("Etablissement", "Du", "Verse", "Reste"):
                row.cell(h, style=FontFace(emphasis="BOLD"))
            for r in retards:
                row = table.row()
                row.cell(pdf_safe(r["etablissement"].nom))
                row.cell(money(r["montant_du"]))
                row.cell(money(r["montant_paye"]))
                row.cell(money(r["reste"]))

    return bytes(pdf.output())


def generate_annual_report_pdf(
    db: Session, annee_scolaire: str, date_debut: datetime.date, date_fin: datetime.date, generated_by: str
) -> bytes:
    """Rapport d'activites annuel consolide (reunions, activites, finances, cartes,
    delegations) destine a l'Assemblee Generale."""
    finances = _period_data(db, date_debut, date_fin)
    solde_cumule = _solde_cumule_au(db, date_fin)
    retards = etablissements_en_retard(db, annee_scolaire)

    reunions = (
        db.query(models.Reunion)
        .filter(
            models.Reunion.delegation_id.is_(None),
            models.Reunion.date >= date_debut,
            models.Reunion.date <= date_fin,
        )
        .order_by(models.Reunion.date)
        .all()
    )
    activites = (
        db.query(models.Activity)
        .filter(models.Activity.event_date >= date_debut, models.Activity.event_date <= date_fin)
        .order_by(models.Activity.event_date)
        .all()
    )
    cartes_emises = (
        db.query(models.CarteMembre)
        .filter(
            models.CarteMembre.status == "disponible",
            models.CarteMembre.date_disponibilite.isnot(None),
            models.CarteMembre.date_disponibilite >= datetime.datetime.combine(date_debut, datetime.time.min),
            models.CarteMembre.date_disponibilite <= datetime.datetime.combine(date_fin, datetime.time.max),
        )
        .count()
    )
    delegations = db.query(models.Delegation).order_by(models.Delegation.nom).all()
    etablissements_total = db.query(models.Etablissement).count()

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()
    _add_logo(pdf)

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(11, 61, 46)
    pdf.cell(0, 10, "LECIM - Rapport d'activites annuel", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 7, f"Annee scolaire {annee_scolaire} (Assemblee Generale)", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(
        0, 7, f"Genere le {datetime.date.today().strftime('%d/%m/%Y')} par {generated_by}",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.ln(4)

    def section_title(text: str) -> None:
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(11, 61, 46)
        pdf.cell(0, 9, text, new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(30, 30, 30)

    def kv_row(label: str, value: str, bold: bool = False) -> None:
        pdf.set_font("Helvetica", "B" if bold else "", 11)
        pdf.cell(100, 7, label)
        pdf.cell(0, 7, value, new_x="LMARGIN", new_y="NEXT")

    section_title("Vue d'ensemble")
    kv_row("Reunions du Bureau Executif National tenues", str(len(reunions)))
    kv_row("Activites organisees", str(len(activites)))
    kv_row("Cartes de membres remises", str(cartes_emises))
    kv_row("Etablissements affilies", str(etablissements_total))
    kv_row("Delegations regionales actives", str(len(delegations)))
    pdf.ln(6)

    section_title("Bilan financier de l'annee")
    kv_row("Droits d'adhesion encaisses", money(finances["total_adhesions"]))
    kv_row("Cotisations encaissees", money(finances["total_cotisations"]))
    kv_row("Autres recettes", money(finances["total_recettes"]))
    kv_row("Ventes de livres", money(finances["total_ventes_livres"]))
    kv_row("Droits d'examens", money(finances["total_droits_examens"]))
    kv_row("Total des entrees", money(finances["total_entrees"]), bold=True)
    kv_row("Depenses", money(finances["total_depenses"]))
    kv_row("Solde de l'annee", money(finances["solde_periode"]), bold=True)
    kv_row("Solde cumule au " + date_fin.strftime("%d/%m/%Y"), money(solde_cumule), bold=True)
    pdf.ln(4)

    _draw_bar_chart(
        pdf,
        pdf.l_margin,
        pdf.get_y(),
        pdf.w - pdf.l_margin - pdf.r_margin,
        40,
        [
            ("Entrees", finances["total_entrees"], GREEN),
            ("Depenses", finances["total_depenses"], GOLD),
            ("Solde", finances["solde_periode"], GREEN if finances["solde_periode"] >= 0 else (192, 57, 43)),
        ],
    )
    pdf.ln(8)

    if reunions:
        section_title("Reunions du Bureau Executif National")
        pdf.set_font("Helvetica", "", 10)
        with pdf.table(col_widths=(35, 90, 40, 25), text_align=("LEFT", "LEFT", "LEFT", "RIGHT")) as table:
            row = table.row()
            for h in ("Date", "Titre", "Lieu", "Presents"):
                row.cell(h, style=FontFace(emphasis="BOLD"))
            for r in reunions:
                row = table.row()
                row.cell(r.date.strftime("%d/%m/%Y"))
                row.cell(pdf_safe(r.title))
                row.cell(pdf_safe(r.lieu) or "-")
                row.cell(f"{r.present_count}/{len(r.presences)}")
        pdf.ln(6)

    if activites:
        section_title("Activites organisees")
        pdf.set_font("Helvetica", "", 10)
        with pdf.table(col_widths=(35, 150) , text_align=("LEFT", "LEFT")) as table:
            row = table.row()
            for h in ("Date", "Titre"):
                row.cell(h, style=FontFace(emphasis="BOLD"))
            for a in activites:
                row = table.row()
                row.cell(a.event_date.strftime("%d/%m/%Y"))
                row.cell(pdf_safe(a.title))
        pdf.ln(6)

    if delegations:
        section_title("Delegations regionales")
        pdf.set_font("Helvetica", "", 10)
        with pdf.table(col_widths=(95, 90), text_align=("LEFT", "LEFT")) as table:
            row = table.row()
            for h in ("Delegation", "Region"):
                row.cell(h, style=FontFace(emphasis="BOLD"))
            for d in delegations:
                row = table.row()
                row.cell(pdf_safe(d.nom))
                row.cell(pdf_safe(d.region) or "-")
        pdf.ln(6)

    if retards:
        section_title(f"Etablissements en retard de cotisation ({annee_scolaire})")
        pdf.set_font("Helvetica", "", 10)
        with pdf.table(col_widths=(70, 40, 40, 40), text_align=("LEFT", "RIGHT", "RIGHT", "RIGHT")) as table:
            row = table.row()
            for h in ("Etablissement", "Du", "Verse", "Reste"):
                row.cell(h, style=FontFace(emphasis="BOLD"))
            for r in retards:
                row = table.row()
                row.cell(pdf_safe(r["etablissement"].nom))
                row.cell(money(r["montant_du"]))
                row.cell(money(r["montant_paye"]))
                row.cell(money(r["reste"]))

    return bytes(pdf.output())


def _variation(a: int, b: int) -> str:
    diff = b - a
    sign = "+" if diff >= 0 else ""
    if a == 0:
        pct = "-" if b == 0 else "+100%" if b > 0 else "-100%"
    else:
        pct = f"{sign}{(diff / abs(a)) * 100:.0f}%"
    return f"{sign}{money(diff)} ({pct})"


def generate_comparative_report_pdf(
    db: Session,
    date_debut_a: datetime.date,
    date_fin_a: datetime.date,
    date_debut_b: datetime.date,
    date_fin_b: datetime.date,
    generated_by: str,
    label_a: str = "Periode A",
    label_b: str = "Periode B",
) -> bytes:
    data_a = _period_data(db, date_debut_a, date_fin_a)
    data_b = _period_data(db, date_debut_b, date_fin_b)

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()
    _add_logo(pdf)

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(11, 61, 46)
    pdf.cell(0, 10, "LECIM - Rapport financier comparatif", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(
        0, 7,
        f"{label_a} : {date_debut_a.strftime('%d/%m/%Y')} au {date_fin_a.strftime('%d/%m/%Y')}",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.cell(
        0, 7,
        f"{label_b} : {date_debut_b.strftime('%d/%m/%Y')} au {date_fin_b.strftime('%d/%m/%Y')}",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.cell(
        0, 7, f"Genere le {datetime.date.today().strftime('%d/%m/%Y')} par {generated_by}",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.ln(4)

    def section_title(text: str) -> None:
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(11, 61, 46)
        pdf.cell(0, 9, text, new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(30, 30, 30)

    section_title("Comparaison des totaux")
    pdf.set_font("Helvetica", "", 10)
    rows = [
        ("Droits d'adhesion", data_a["total_adhesions"], data_b["total_adhesions"]),
        ("Cotisations", data_a["total_cotisations"], data_b["total_cotisations"]),
        ("Autres recettes", data_a["total_recettes"], data_b["total_recettes"]),
        ("Ventes de livres", data_a["total_ventes_livres"], data_b["total_ventes_livres"]),
        ("Droits d'examens", data_a["total_droits_examens"], data_b["total_droits_examens"]),
        ("Total des entrees", data_a["total_entrees"], data_b["total_entrees"]),
        ("Depenses", data_a["total_depenses"], data_b["total_depenses"]),
        ("Solde de la periode", data_a["solde_periode"], data_b["solde_periode"]),
    ]
    with pdf.table(col_widths=(60, 40, 40, 55), text_align=("LEFT", "RIGHT", "RIGHT", "RIGHT")) as table:
        row = table.row()
        for h in ("Indicateur", label_a, label_b, "Ecart"):
            row.cell(h, style=FontFace(emphasis="BOLD"))
        for label, val_a, val_b in rows:
            row = table.row()
            row.cell(label)
            row.cell(money(val_a))
            row.cell(money(val_b))
            row.cell(_variation(val_a, val_b))
    pdf.ln(10)

    section_title("Entrees / Depenses / Solde")
    _draw_grouped_bar_chart(
        pdf,
        pdf.l_margin,
        pdf.get_y(),
        pdf.w - pdf.l_margin - pdf.r_margin,
        42,
        ["Entrees", "Depenses", "Solde"],
        [data_a["total_entrees"], data_a["total_depenses"], data_a["solde_periode"]],
        [data_b["total_entrees"], data_b["total_depenses"], data_b["solde_periode"]],
        label_a,
        label_b,
    )
    pdf.ln(6)

    categories = sorted(set(data_a["recettes_par_categorie"]) | set(data_b["recettes_par_categorie"]))
    if categories:
        section_title("Autres recettes par categorie")
        serie_a = [data_a["recettes_par_categorie"].get(c, 0) for c in categories]
        serie_b = [data_b["recettes_par_categorie"].get(c, 0) for c in categories]
        _draw_grouped_bar_chart(
            pdf,
            pdf.l_margin,
            pdf.get_y(),
            pdf.w - pdf.l_margin - pdf.r_margin,
            42,
            [RECETTE_CATEGORIES.get(c, c)[:12] for c in categories],
            serie_a,
            serie_b,
            label_a,
            label_b,
        )
        pdf.ln(6)

    return bytes(pdf.output())


def export_transactions_csv(db: Session, date_debut: datetime.date, date_fin: datetime.date) -> str:
    data = _period_data(db, date_debut, date_fin)
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow(["Date", "Type", "Categorie", "Libelle / Etablissement", "Montant (FCFA)"])

    for a in data["adhesions"]:
        writer.writerow([a.date_paiement.isoformat(), "Adhesion", "-", a.etablissement.nom, a.montant])
    for c in data["cotisations"]:
        writer.writerow(
            [c.date_paiement.isoformat(), "Cotisation", c.annee_scolaire, c.etablissement.nom, c.montant_paye]
        )
    for r in data["recettes"]:
        writer.writerow(
            [r.date.isoformat(), "Recette", RECETTE_CATEGORIES.get(r.categorie, r.categorie), r.libelle, r.montant]
        )
    for v in data["ventes_livres"]:
        writer.writerow([v.date.isoformat(), "Vente de livre", f"Qte {v.quantite}", v.titre, v.montant])
    for de in data["droits_examens"]:
        writer.writerow([
            de.date.isoformat(), "Droit d'examen", de.type_examen or "-",
            de.libelle + (f" ({de.etablissement.nom})" if de.etablissement else ""), de.montant,
        ])
    for d in data["depenses"]:
        writer.writerow([d.date.isoformat(), "Depense", "-", d.libelle, -d.montant])

    return buffer.getvalue()


def export_resultats_csv(db: Session) -> str:
    items = (
        db.query(models.ResultatExamen)
        .options(joinedload(models.ResultatExamen.etablissement))
        .order_by(models.ResultatExamen.annee_scolaire.desc(), models.ResultatExamen.id.desc())
        .all()
    )
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow(["Annee scolaire", "Etablissement", "Examen", "Inscrits", "Admis", "Taux de reussite (%)", "Publie"])
    for r in items:
        writer.writerow([
            r.annee_scolaire, r.etablissement.nom, r.type_examen,
            r.nombre_inscrits, r.nombre_admis, r.taux_reussite, "Oui" if r.is_published else "Non",
        ])
    return buffer.getvalue()


def export_effectifs_csv(db: Session) -> str:
    items = (
        db.query(models.Effectif)
        .options(joinedload(models.Effectif.etablissement))
        .order_by(models.Effectif.annee_scolaire.desc(), models.Effectif.niveau)
        .all()
    )
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow(["Annee scolaire", "Etablissement", "Niveau", "Garcons", "Filles", "Total"])
    for e in items:
        writer.writerow([e.annee_scolaire, e.etablissement.nom, e.niveau_label, e.nombre_garcons, e.nombre_filles, e.total])
    return buffer.getvalue()


def export_transactions_xlsx(db: Session, date_debut: datetime.date, date_fin: datetime.date) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    data = _period_data(db, date_debut, date_fin)

    wb = Workbook()

    header_fill = PatternFill(start_color="0B3D2E", end_color="0B3D2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="0B3D2E", bold=True, size=14)
    bold_font = Font(bold=True)

    # ---------- Feuille Résumé ----------
    resume = wb.active
    resume.title = "Résumé"
    resume["A1"] = "LECIM — Rapport financier"
    resume["A1"].font = title_font
    resume["A2"] = f"Période du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"

    if LOGO_PATH.exists():
        from openpyxl.drawing.image import Image as XLImage

        xl_logo = XLImage(str(LOGO_PATH))
        xl_logo.width = 60
        xl_logo.height = 60
        resume.add_image(xl_logo, "D1")

    resume_rows = [
        ("Droits d'adhésion encaissés", data["total_adhesions"]),
        ("Cotisations encaissées", data["total_cotisations"]),
        ("Autres recettes", data["total_recettes"]),
        ("Ventes de livres", data["total_ventes_livres"]),
        ("Droits d'examens", data["total_droits_examens"]),
        ("Total des entrées", data["total_entrees"]),
        ("Dépenses", data["total_depenses"]),
        ("Solde de la période", data["solde_periode"]),
    ]
    resume["A4"] = "Indicateur"
    resume["B4"] = "Montant (FCFA)"
    for cell in ("A4", "B4"):
        resume[cell].fill = header_fill
        resume[cell].font = header_font
    for i, (label, value) in enumerate(resume_rows, start=5):
        resume[f"A{i}"] = label
        resume[f"B{i}"] = value
        resume[f"B{i}"].number_format = "#,##0"
        if label in ("Total des entrées", "Solde de la période"):
            resume[f"A{i}"].font = bold_font
            resume[f"B{i}"].font = bold_font
    resume.column_dimensions["A"].width = 32
    resume.column_dimensions["B"].width = 18

    if data["recettes_par_categorie"]:
        start_row = 5 + len(resume_rows) + 2
        resume[f"A{start_row}"] = "Recettes par catégorie"
        resume[f"A{start_row}"].font = bold_font
        resume[f"A{start_row + 1}"] = "Catégorie"
        resume[f"B{start_row + 1}"] = "Montant (FCFA)"
        for cell in (f"A{start_row + 1}", f"B{start_row + 1}"):
            resume[cell].fill = header_fill
            resume[cell].font = header_font
        row = start_row + 2
        for categorie, montant in data["recettes_par_categorie"].items():
            resume[f"A{row}"] = RECETTE_CATEGORIES.get(categorie, categorie)
            resume[f"B{row}"] = montant
            resume[f"B{row}"].number_format = "#,##0"
            row += 1

        chart = BarChart()
        chart.title = "Recettes par catégorie"
        chart.y_axis.title = "FCFA"
        chart.style = 10
        data_ref = Reference(resume, min_col=2, min_row=start_row + 1, max_row=row - 1)
        cats_ref = Reference(resume, min_col=1, min_row=start_row + 2, max_row=row - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 16
        chart.height = 9
        resume.add_chart(chart, f"D{start_row}")

    # ---------- Feuille Transactions ----------
    sheet = wb.create_sheet("Transactions")
    headers = ["Date", "Type", "Catégorie", "Libellé / Établissement", "Montant (FCFA)"]
    sheet.append(headers)
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for a in data["adhesions"]:
        sheet.append([a.date_paiement, "Adhésion", "-", a.etablissement.nom, a.montant])
    for c in data["cotisations"]:
        sheet.append([c.date_paiement, "Cotisation", c.annee_scolaire, c.etablissement.nom, c.montant_paye])
    for r in data["recettes"]:
        sheet.append(
            [r.date, "Recette", RECETTE_CATEGORIES.get(r.categorie, r.categorie), r.libelle, r.montant]
        )
    for v in data["ventes_livres"]:
        sheet.append([v.date, "Vente de livre", f"Qté {v.quantite}", v.titre, v.montant])
    for de in data["droits_examens"]:
        sheet.append([
            de.date, "Droit d'examen", de.type_examen or "-",
            de.libelle + (f" ({de.etablissement.nom})" if de.etablissement else ""), de.montant,
        ])
    for d in data["depenses"]:
        sheet.append([d.date, "Dépense", "-", d.libelle, -d.montant])

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        row[0].number_format = "DD/MM/YYYY"
    for row in sheet.iter_rows(min_row=2, min_col=5, max_col=5):
        row[0].number_format = "#,##0"

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 22
    sheet.column_dimensions["D"].width = 38
    sheet.column_dimensions["E"].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
