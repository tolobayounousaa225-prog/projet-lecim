"""Importe les établissements affiliés depuis le fichier Excel fourni par le BEN
(colonnes : No, CODE ADH, NOM DE L'ECOLE, LOCALITE, FIXE, PORTABLE, DATE ADH,
ETAT, AGREE, Enregistré sous le No). N'écrase rien : ignore les codes d'adhésion
déjà présents en base pour permettre de relancer le script sans dupliquer."""

import datetime
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from app.database import SessionLocal
from app import models

EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else None
if not EXCEL_PATH:
    print("Usage: import_ecoles_excel.py <chemin_du_fichier.xlsx>")
    sys.exit(1)

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb["ورقة1"]

db = SessionLocal()
existing_codes = {c for (c,) in db.query(models.Etablissement.code_adhesion).filter(models.Etablissement.code_adhesion.isnot(None))}

created = 0
skipped = 0
for r in range(4, ws.max_row + 1):
    no, code, nom, localite, fixe, portable, date_adh, etat, agree, num, prefix = [
        ws.cell(row=r, column=c).value for c in range(1, 12)
    ]
    if not nom:
        continue
    if code in existing_codes:
        skipped += 1
        continue

    tel = None
    for cand in (portable, fixe):
        if cand and str(cand).strip() not in ("0", ""):
            tel = str(cand).strip()
            break

    localite_clean = None
    if localite and "?" not in str(localite):
        localite_clean = str(localite).strip()

    date_adhesion = date_adh.date() if isinstance(date_adh, datetime.datetime) else None

    etab = models.Etablissement(
        code_adhesion=code,
        nom=str(nom).strip(),
        bureau_local=localite_clean,
        contact_telephone=tel,
        date_adhesion=date_adhesion,
        statut="non_subventionne",
        type_enseignement="les_deux",
    )
    db.add(etab)
    created += 1

db.commit()
print(f"Créés : {created} — ignorés (code déjà présent) : {skipped}")
db.close()
